import openpyxl
from openpyxl import Workbook

# Create src/test/resources directory if it doesn't exist
import os
os.makedirs("src/test/resources", exist_ok=True)

# Create test_data.xlsx
workbook_test_data = Workbook()
sheet1 = workbook_test_data.active
sheet1.title = "Sheet1"
sheet1.append(["Name", "Age", "City"])
sheet1.append(["Alice", 30, "New York"])
sheet1.append(["Bob", 24, "London"])

sheet2 = workbook_test_data.create_sheet("Sheet2")
sheet2.append(["Product", "Price", "Stock"])
sheet2.append(["Apple", 1.0, 100])
sheet2.append(["Banana", 0.5, 150])

workbook_test_data.save("src/test/resources/test_data.xlsx")

# Create empty_test.xlsx
workbook_empty = Workbook()
# A new workbook always has one sheet.
workbook_empty.save("src/test/resources/empty_test.xlsx")

# Create cell_types_test.xlsx
workbook_cell_types = Workbook()
sheet_cell_types = workbook_cell_types.active
sheet_cell_types.title = "CellTypes"
sheet_cell_types.append(["StringVal", "NumericVal", "BooleanVal", "FormulaVal"])
# For the formula, Excel typically stores the last calculated value.
# openpyxl will write the formula, but the value read by POI will be the cached value.
# If the file is opened and saved by Excel, "=A2" where A2 is "Test" will result in "Test".
# For a direct openpyxl write, we can set the value directly.
cell = sheet_cell_types.cell(row=2, column=4)
cell.value = "Test" # Setting the expected result of the formula directly
# If we wanted to write an actual formula: cell.value = "=A2" 
# but then reading it without Excel evaluating it might give 0 or an error depending on POI's formula evaluation capabilities.
# For this test, we'll assume the formula's *cached value* is what we want to test.
sheet_cell_types.append(["Test", 123.45, True, cell.value])


workbook_cell_types.save("src/test/resources/cell_types_test.xlsx")

print("Excel files created successfully in src/test/resources/")
